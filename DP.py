import customtkinter as ctk
import openpyxl, pathlib, phonenumbers, operator, re
from openpyxl import Workbook
from openpyxl.styles import Alignment
from tkinter import END, messagebox
from phonenumbers import PhoneNumberFormat

class BackEnd():
    
    def RH(self):
        self.ficheiro = pathlib.Path(r'Planilha admissional.xlsx')
        if self.ficheiro.exists():
            pass
        else:
            self.ficheiro = Workbook()
            self.folha = self.ficheiro.active
            self.folha.title = 'Dados de funcionários'
            self.ficheiro.save(r'Planilha admissional.xlsx')

            self.folha['A1'] = 'Ficha'
            self.folha['B1'] = 'Sap'
            self.folha['C1'] = 'Funcionário'
            self.folha['D1'] = 'Cargo'
            self.folha['E1'] = 'Departamento'
            self.folha['F1'] = 'Data de Nasc'
            self.folha['G1'] = 'Contato'
            self.folha['H1'] = 'RG'
            self.folha['I1'] = 'CPF'
            self.folha['J1'] = 'CTPS'
            self.folha['K1'] = 'SÉRIE'
            self.folha['L1'] = 'PIS'
            self.folha['M1'] = 'Dependentes'
            self.folha['N1'] = 'Sexo'
            self.folha['O1'] = 'Estado Civil'
            self.folha['P1'] = 'Data de Admissão'
            self.folha['Q1'] = 'Salário'
            self.folha['R1'] = 'Escolaridade'

            self.ficheiro.save(r'Planilha admissional.xlsx')
  
    def Salvar_informações(self):

        Colunas_Centralizar = [1, 2 , 3, 
                               4 , 5, 6, 
                               7, 8, 9, 
                               10, 11, 12, 
                               13, 14, 15, 
                               16, 17, 18]
        
        def formatar_nome(nome):
                
            palavras = nome.split()  
            nome_formatado = []

            for palavra in palavras:
                if len(palavra) > 2:
                    nome_formatado.append(palavra.capitalize())  
                else:
                    nome_formatado.append(palavra.lower())  

            return ' '.join(nome_formatado)  

        self.ficha = (self.Ficha_Entry.get())
        self.sap = (self.Sap_Entry.get())
        self.nome = formatar_nome(self.Nome_Entry.get())
        self.cargo = self.Cargo_Menu.get()
        self.departamento = self.Departamento_Menu.get()
        self.escolaridade = self.Escolaridade_Menu.get()
        self.estado_civil = self.Estado_Civil_Menu.get()
        self.sexo = self.Sexo_Menu.get()
        self.data_admissional = self.Data_Admissional_Entry.get()
        self.data_nasc = self.Data_Nasc_Entry.get()
        self.dependentes = (self.Dependentes_Entry.get())
        self.salario = self.Salario_Entry.get()
        self.contato = (self.Contato_Entry.get())
        self.rg = self.RG_Entry.get()
        self.cpf = self.CPF_Entry.get()
        self.ctps = self.CTPS_Entry.get()
        self.serie = self.SERIE_Entry.get()
        self.pis = self.PIS_Entry.get()

        if (not self.ficha or not self.sap or not self.nome or 
            not self.cargo or not self.departamento or not self.escolaridade or 
            not self.estado_civil or not self.sexo or not self.data_admissional or not self.data_nasc or 
            not self.dependentes or not self.salario or not self.contato or not self.rg or not self.cpf or 
            not self.ctps or not self.serie or not self.pis):
            
            self.ficheiro = openpyxl.load_workbook(r'Planilha admissional.xlsx')
            self.folha = self.ficheiro.get_sheet_by_name(r'Dados de funcionários')

            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nPor favor, preencha todos os campos :)')
            return

        self.ficheiro = openpyxl.load_workbook(r'Planilha admissional.xlsx')
        self.folha = self.ficheiro.get_sheet_by_name(r'Dados de funcionários')
    
        if not self.nome.replace(' ', '').isalpha():
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nO campo Nome, deve conter apenas letras :)')
            return
        
        try:
            int(self.Ficha_Entry.get())
            int(self.Sap_Entry.get())
            int(self.Dependentes_Entry.get())
            int(self.CTPS_Entry.get())
            
        except ValueError:
                messagebox.showerror(title = 'Sistema Admissional', 
                                     message = 'ERRO :/\nOs campos de Ficha, Sap, Dependentes e CTPS, devem conter apenas números :)')
                return
        
        salario_formatado = self.Salario_Entry.get()
        if not re.match(r'^\d*(,\d*)?$', salario_formatado):
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nO campo Salario, deve conter apenas números, ou somente uma única vírgula :)')
            return

        self.salario = float(salario_formatado.replace(',', '.'))
        salario_formatado = self.Salario_Entry.get()
        if not re.match(r'^\d*(,\d{0,1}\d)?$', salario_formatado):
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nNo campo Salario, digite a vírgula somente para adicionar centavos :)')
            return

        self.salario = float(salario_formatado.replace(',', '.'))

        self.rg = self.RG_Entry.get()
        self.serie = self.SERIE_Entry.get()
        if not re.match(r'^[0-9A-Z-]+$', self.rg) or not re.match(r'^[0-9A-Z-]+$', self.serie):
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nOs campos RG e SERIE, devem conter apenas números, letras maiúsculas e o caractere - :)')
            return
        
        self.pis = self.PIS_Entry.get()
        if not re.match(r'^[0-9.-]+$', self.pis):
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nO campo PIS, deve conter apenas números, pontos e o caractere - :)')
            return
        
        cpf_valido = re.match(r'^[\d.-]+$', self.cpf)
        if not cpf_valido:
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nO campo CPF, deve conter apenas números :)')
            return
        
        if len(re.sub(r'\D', '', self.cpf)) != 11:
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message ='ERRO :/\nO campo CPF, deve conter exatamente 11 números :)')
            return

        cpf_formato = re.compile(r'^\d{3}\.\d{3}\.\d{3}-\d{2}$')
        if not cpf_formato.match(self.cpf):
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nO campo CPF, deve ter o formato 000.000.000-00 :)')
            return
        
        if not self.validar_data(self.Data_Admissional_Entry.get()) or not self.validar_data(self.Data_Nasc_Entry.get()):
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nAs datas devem estar no formato:\n00/00/0000, 00/ 00/ 0000 ou 00 / 00 / 0000 :)')
            return
        
        if not re.match(r'^[\d() -]+$', self.Contato_Entry.get()):
            messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nO campo Contato, deve conter apenas números :)')
            return
        
        self.Contato_BR = self.Contato_Entry.get()
        Contato_Formatado = r'^\(\d{2}\)\s?\d{5}-\d{4}$'   
                                                
        if len(re.sub(r'\D', '', self.Contato_BR)) != 11:
                    messagebox.showerror(title = 'Sistema Admissional', 
                                 message = 'ERRO :/\nO número de Contato, deve conter exatamente 11 números :)')
                    return
        
        if not re.match(Contato_Formatado, self.Contato_BR):
                messagebox.showerror(title = 'Sistema Admissional', 
                                     message = 'ERRO :/\nSalve o Contato com esse formato: (88) 98888-8888 :)')
                return
        
        if not re.match(r'\(\d{2}\)\s?9', self.Contato_BR):
                messagebox.showerror(title = 'Sistema Admissional', 
                                     message = 'ERRO :/\nPor favor, insira o número 9 no Contato após o DDD :)')
                return
        
        DDD_BR = ['61', '62', '64', '65', '66', '67', '82', '71', '73', '74', '75', '77', 
                  '85', '88', '98', '99', '83', '81', '87', '86', '89', '84', '79', '68','96', 
                  '92', '97', '91', '93', '94', '69', '95', '63', '27', '28', '31', '32', '33', 
                  '34', '35', '37', '38', '21', '22', '24', '11', '12', '13', '14', '15', '16', '17', 
                  '18', '19', '41', '42', '43', '44', '45','46', '51', '53', '54', '55', '47', '48', '49']

        self.contato_ajustado = phonenumbers.parse(self.Contato_BR, 'BR')
        self.contato_formatado = phonenumbers.format_number(self.contato_ajustado, PhoneNumberFormat.NATIONAL)

        DDD_BR_Formatado = re.search(r'\((\d{2})\)', self.contato_formatado).group(1)
        if DDD_BR_Formatado not in DDD_BR:
            messagebox.showerror(title = 'Sistema Admissional',
                                 message = 'ERRO:/\nO DDD do contato não é do Brasil (BR)\nPor favor, insira um DDD válido :)')
            return

        self.ficheiro = openpyxl.load_workbook(r'Planilha admissional.xlsx')
        self.folha = self.ficheiro.get_sheet_by_name(r'Dados de funcionários')

        row_to_insert = 2  
        while self.folha.cell(column = 1, row = row_to_insert).value:
            row_to_insert += 1

        self.folha.cell(column = 1, row = row_to_insert, value = self.ficha)
        self.folha.cell(column = 2, row = row_to_insert, value = self.sap)
        self.folha.cell(column = 3, row = row_to_insert, value = self.nome)
        self.folha.cell(column = 4, row = row_to_insert, value = self.cargo)
        self.folha.cell(column = 5, row = row_to_insert, value = self.departamento)
        self.folha.cell(column = 6, row = row_to_insert, value = self.data_nasc)
        self.folha.cell(column = 7, row = row_to_insert, value = self.contato_formatado)
        self.folha.cell(column = 7, row = row_to_insert).number_format = '0'
        self.folha.cell(column = 8, row = row_to_insert, value = self.rg)
        self.folha.cell(column = 9, row = row_to_insert, value = self.cpf)
        self.folha.cell(column = 10, row = row_to_insert, value = self.ctps)
        self.folha.cell(column = 11, row = row_to_insert, value = self.serie)
        self.folha.cell(column = 12, row = row_to_insert, value = self.pis)
        self.folha.cell(column = 13, row = row_to_insert, value = self.dependentes)
        self.folha.cell(column = 14, row = row_to_insert, value = self.sexo)
        self.folha.cell(column = 15, row = row_to_insert, value = self.estado_civil)
        self.folha.cell(column = 16, row = row_to_insert, value = self.data_admissional)
        self.folha.cell(column = 17, row = row_to_insert, value = float(self.salario))
        self.folha.cell(column = 17, row = row_to_insert).number_format = 'R$ #,##0.00'
        self.folha.cell(column = 18, row = row_to_insert, value = self.escolaridade)
   
        for Coluna in Colunas_Centralizar:
                for Linha in range(2, self.folha.max_row + 1):  
                    celula = self.folha.cell(column = Coluna, row = Linha)
                    celula.alignment = Alignment(horizontal = 'center')

        self.ficheiro.save(r'Planilha admissional.xlsx')
        print('Dados salvos com sucesso!')
        
        messagebox.showinfo(title = 'Sistema Admissional', 
                            message = 'Dados salvos com sucesso!')
        
    def Salvar_e_ordenar_dados(self):

        funcionarios = []

        for row in self.folha.iter_rows(min_row = 2, values_only = True):
            
            (ficha, sap, nome, cargo, 
             departamento, data_nasc, contato, 
             rg, cpf, ctps, serie, pis, dependentes, 
             sexo, estado_civil, data_admissional, 
             salario, escolaridade) = row
            
            if nome:

                funcionarios.append({'ficha': int(ficha), 'sap': int(sap), 'nome': nome, 
                                     'cargo': cargo, 'departamento': departamento, 'data_nasc': data_nasc, 
                                     'contato': contato, 'rg': rg, 'cpf': cpf, 
                                     'ctps': int(ctps), 'serie': serie, 'pis': pis, 
                                     'dependentes': int(dependentes), 'sexo': sexo, 'estado_civil': estado_civil, 
                                     'data_admissional': data_admissional, 'salario': salario, 
                                     'escolaridade': escolaridade})

        funcionarios.sort(key = operator.itemgetter('nome'))

        for row_index, funcionario in enumerate(funcionarios, start = 2):
            self.folha.cell(row = row_index, column = 1, value = funcionario['ficha'])
            self.folha.cell(row = row_index, column = 2, value = funcionario['sap'])
            self.folha.cell(row = row_index, column = 3, value = funcionario['nome'])
            self.folha.cell(row = row_index, column = 4, value = funcionario['cargo'])
            self.folha.cell(row = row_index, column = 5, value = funcionario['departamento'])
            self.folha.cell(row = row_index, column = 6, value = funcionario['data_nasc'])
            self.folha.cell(row = row_index, column = 7, value = funcionario['contato'])
            self.folha.cell(row = row_index, column = 8, value = funcionario['rg'])
            self.folha.cell(row = row_index, column = 9, value = funcionario['cpf'])
            self.folha.cell(row = row_index, column = 10, value = funcionario['ctps'])
            self.folha.cell(row = row_index, column = 11, value = funcionario['serie'])
            self.folha.cell(row = row_index, column = 12, value = funcionario['pis'])
            self.folha.cell(row = row_index, column = 13, value = funcionario['dependentes'])
            self.folha.cell(row = row_index, column = 14, value = funcionario['sexo'])
            self.folha.cell(row = row_index, column = 15, value = funcionario['estado_civil'])
            self.folha.cell(row = row_index, column = 16, value = funcionario['data_admissional'])
            self.folha.cell(row = row_index, column = 17, value = funcionario['salario'])
            self.folha.cell(row = row_index, column = 18, value = funcionario['escolaridade'])
        
        self.ficheiro.save(r'Planilha admissional.xlsx')

class App(ctk.CTk, BackEnd):
    def __init__(self):
        super().__init__()
        self.Janela()
        self.Informação_Pessoal()
        self.RH()
        self.Tema()
        self.Salvar_Dados()
        self.Limpar_Informaçoes()
        
    def Janela(self):
        self.title('Departamento Pessoal')
        self.geometry('660x490')
        self.minsize(width = 660, height = 490)
        
        ctk.CTkLabel(self, text = 'Sistema Admissional', 
                     font = ('arialbold', 20), 
                     width = 1300).pack()
        
    def Tema(self):

        self.switch_var = ctk.StringVar(value = 'on')
        ctk.set_appearance_mode('light')

        def Modo():
           
            if self.switch_var.get() == 'Desativado':
                ctk.set_appearance_mode('light')

            elif self.switch_var.get() == 'Ativado':
                ctk.set_appearance_mode('dark')

            print('O switch esta: ', self.switch_var.get())

        self.switch = ctk.CTkSwitch(self,  
                                     text = 'Tema',
                                     command = Modo,
                                     variable = self.switch_var,
                                     onvalue = 'Ativado',
                                     offvalue = 'Desativado', 
                                     font = ('arialbold', 14), 
                                     switch_width = 30 , 
                                     switch_height = 15)
        
        self.switch.place(x = 10, y = 0)
        
    def Informação_Pessoal(self):

        self.tabview = ctk.CTkTabview(self, 
                                      width = 550, 
                                      height = 415, 
                                      corner_radius = 0, 
                                      border_width = 1)

        self.tabview.pack()
        
        self.Ficha_Label = ctk.CTkLabel(self.tabview, 
                                        text = 'Ficha', 
                                        font = ('arialbold', 14))
        
        self.Ficha_Entry = ctk.CTkEntry(self.tabview, 
                                        width = 60, 
                                        height = 1, 
                                        border_width = 1, 
                                        corner_radius = 0, 
                                        font = ('arialbold', 14))
        
        self.Sap_Label = ctk.CTkLabel(self.tabview, 
                                      text = 'Sap', 
                                      font = ('arialbold', 14))
        
        self.Sap_Entry = ctk.CTkEntry(self.tabview, 
                                      width = 60, 
                                      height = 1, 
                                      border_width = 1, 
                                      corner_radius = 0, 
                                      font = ('arialbold', 14))
        
        self.Nome_Label = ctk.CTkLabel(self.tabview, 
                                       text = 'Nome', 
                                       font = ('arialbold', 14))
        
        self.Nome_Entry = ctk.CTkEntry(self.tabview, 
                                       placeholder_text = 'Por favor, digite o nome completo',
                                       width = 250, 
                                       height = 1, 
                                       border_width = 1, 
                                       corner_radius = 0, 
                                       font = ('arialbold', 14))
        
        self.Data_Admissional_Label = ctk.CTkLabel(self.tabview, 
                                                   text = 'Data de Admissao', 
                                                   font = ('arialbold', 14))
        
        self.Data_Admissional_Entry = ctk.CTkEntry(self.tabview, 
                                                   placeholder_text = '00 / 00 / 0000', 
                                                   width = 130,
                                                   height = 1, 
                                                   border_width = 1, 
                                                   corner_radius = 0, 
                                                   font = ('arialbold', 14))
        
        self.Data_Nasc_Label = ctk.CTkLabel(self.tabview, 
                                            text = 'Data de Nascimento', 
                                            font = ('arialbold', 14))
        
        self.Data_Nasc_Entry = ctk.CTkEntry(self.tabview, 
                                            placeholder_text = '00 / 00 / 0000',
                                            width = 130,  
                                            height = 1, 
                                            border_width = 1, 
                                            corner_radius = 0, 
                                            font = ('arialbold', 14))
        
        self.Dependentes_Label = ctk.CTkLabel(self.tabview, 
                                              text = 'Dependentes', 
                                              font = ('arialbold', 14))
        
        self.Dependentes_Entry = ctk.CTkEntry(self.tabview, 
                                              width = 30, 
                                              height = 1, 
                                              border_width = 1, 
                                              corner_radius = 0, 
                                              font = ('arialbold', 14))

        self.Salario_Label = ctk.CTkLabel(self.tabview, 
                                          text = 'Salario', 
                                          font = ('arialbold', 14))
        
        self.Salario_Entry = ctk.CTkEntry(self.tabview, 
                                          width = 80, 
                                          height = 1, 
                                          border_width = 1, 
                                          corner_radius = 0, 
                                          font = ('arialbold', 14))
        
        self.Contato_Label = ctk.CTkLabel(self.tabview, 
                                          text = 'Contato', 
                                          font = ('arialbold', 14))
        
        self.Contato_Entry = ctk.CTkEntry(self.tabview,
                                          placeholder_text = '(88) 98888-8888', 
                                          width = 130, 
                                          height = 1, 
                                          border_width = 1, 
                                          corner_radius = 0, 
                                          font = ('arialbold', 14))
        
        self.RG_Label = ctk.CTkLabel(self.tabview, 
                                     text = 'RG', 
                                     font = ('arialbold', 14))
        
        self.RG_Entry = ctk.CTkEntry(self.tabview, 
                                     width = 250, 
                                     height = 1, 
                                     border_width = 1, 
                                     corner_radius = 0, 
                                     font = ('arialbold', 14))
        
        self.CPF_Label = ctk.CTkLabel(self.tabview, 
                                      text = 'CPF', 
                                      font = ('arialbold', 14))
        
        self.CPF_Entry = ctk.CTkEntry(self.tabview, 
                                      width = 250, 
                                      height = 1, 
                                      border_width = 1, 
                                      corner_radius = 0, 
                                      font = ('arialbold', 14))
        
        self.CTPS_Label = ctk.CTkLabel(self.tabview, 
                                       text = 'CTPS', 
                                       font = ('arialbold', 14))
        
        self.CTPS_Entry = ctk.CTkEntry(self.tabview, 
                                       width = 250, 
                                       height = 1, 
                                       border_width = 1, 
                                       corner_radius = 0, 
                                       font = ('arialbold', 14))
        
        self.SERIE_Label = ctk.CTkLabel(self.tabview, 
                                        text = 'SERIE', 
                                        font = ('arialbold', 14))
        
        self.SERIE_Entry = ctk.CTkEntry(self.tabview, 
                                        width = 250, 
                                        height = 1, 
                                        border_width = 1, 
                                        corner_radius = 0, 
                                        font = ('arialbold', 14))
        
        self.PIS_Label = ctk.CTkLabel(self.tabview, 
                                      text = 'PIS', 
                                      font = ('arialbold', 14))
        
        self.PIS_Entry = ctk.CTkEntry(self.tabview, 
                                      width = 250, 
                                      height = 1, 
                                      border_width = 1, 
                                      corner_radius = 0, 
                                      font = ('arialbold', 14))
        
        def Cargo(Escolha):
            
            print(f'Cargo: {Escolha}')

        self.valor1 = ['Analista de Engenharia', 'Apontador', 
                       'Aprendiz de Aux. Administrativo',
                       'Aux. Administrativo', 'Aux. de Laboratório', 
                       'Aux . de Mecânica', 'Aux. de Serviços Gerais', 
                       'Aux. de Topografia', 'Bandeirinha', 
                       'Carpinteiro', 'Comprador', 
                       'Coordenador Administrativo', 'Contr. de Manutençao I', 
                       'Eletricista de Corrente Continua', 'Enc. de Ar Comprimido', 
                       'Enc. de Departamento Pessoal', 'Enc. de Mecanica', 
                       'Enc. de Obra. de Artes Corrente', 'Enc. de Obra. de Artes Especiais', 
                       'Enc. de Serviços Gerais', 'Enc. de Transportes', 
                       'Eng. de Obra', 'Escavador de Tubulão', 
                       'Feitor de Turma', 'Greidista de Terraplenagem', 
                       'Laboratorista', 'Lavador de Maq. e Veiculos', 
                       'Mec. de Equip. Leves', 'Mec. de Equip. Pesados', 
                       'Mot. de Caminhão Basculante', 'Mot de Caminhão Comboio', 
                       'Mot. de Caminhao Leve', 'Mot. De Caminhao Munck', 
                       'Mot de Caminhão Pipa', 'Mot. de Transporte de Pessoal', 
                       'Médico do Trabalho', 'Op. de Campanula', 
                       'Op. de Escavadeira', 'OP. de Maquinas Pesadas', 
                       'Op. de Mesa Acabadora', 'Op. de Rolo', 
                       'Op. de Retro Escavadeira', 'Op. de Trator Agricola', 
                       'Op. de Trator Esteira', 'Op. de Usina de Concreto', 
                       'Pedreiro', 'Porteiro', 
                       'Servente de Obra', 'Tec. de Segurança do Trabalho', 
                       'Topografo', 'Vigia']
        
        self.Cargo_var = ctk.StringVar(value = 'Cargo')

        self.Cargo_Menu = ctk.CTkOptionMenu(self.tabview, 
                                            values = self.valor1, 
                                            variable = self.Cargo_var, 
                                            command = Cargo, 
                                            width = 250, 
                                            height = 20, 
                                            corner_radius = 0, 
                                            fg_color = ['#fff', '#333'], 
                                            button_color = ['#fff', '#333'], 
                                            button_hover_color = '', 
                                            text_color = ['#333', '#fff'], 
                                            font = ('arialbold', 14))
        
        def Departamento(Escolha):
            
            print(f'Departamento: {Escolha}')

        self.valor2 = ['Administração', 'Almoxarifado', 'Departamento Pessoal', 
                       'Engenharia', 'Mêcanica', 'M.O.D', 'O.A.C', 'O.A.C / Drenagem', 
                       'O.A.E', 'Pavimento Rígido', 'QSMS', 
                       'Serviços Gerais', 'Topografia', 'Vigilancia']
        
        self.Departamento_var = ctk.StringVar(value = 'Departamento')

        self.Departamento_Menu = ctk.CTkOptionMenu(self.tabview, 
                                                   values = self.valor2, 
                                                   variable = self.Departamento_var, 
                                                   command = Departamento, 
                                                   width = 250, 
                                                   height = 20, 
                                                   corner_radius = 0, 
                                                   fg_color = ['#fff', '#333'], 
                                                   button_color = ['#fff', '#333'], 
                                                   button_hover_color = '', 
                                                   text_color = ['#333', '#fff'], 
                                                   font = ('arialbold', 14))
        
        def Escolaridade(Escolha):
            
            print(f'Escolaridade: {Escolha}')

        self.valor3 = ['Alfabetizado', 'Fundamental Completo', 
                       'Fundamental Incompleto', 'Graduação', 
                       'Medio Completo', 'Medio Incompleto', 
                       'Segundo Grau Completo', 'Segundo Grau Incompleto', 
                       'Superior Completo', 'Superior Incompleto']
        
        self.Escolaridade_var = ctk.StringVar(value = 'Escolaridade')

        self.Escolaridade_Menu = ctk.CTkOptionMenu(self.tabview, 
                                                   values = self.valor3, 
                                                   variable = self.Escolaridade_var, 
                                                   command = Escolaridade, 
                                                   width = 250, 
                                                   height = 20, 
                                                   corner_radius = 0, 
                                                   fg_color = ['#fff', '#333'], 
                                                   button_color = ['#fff', '#333'], 
                                                   button_hover_color = '', 
                                                   text_color = ['#333', '#fff'], 
                                                   font = ('arialbold', 14))
        
        def Estado_Civil(Escolha):
            
            print(f'Estado Civil: {Escolha}')

        self.valor4 = ['Solteiro', 'Solteira', 
                       'Casado', 'Casada', 
                       'Divorciado', 'Divorciada', 
                       'Viuvo', 'Viuva', 
                       'Separado', 'Separada']
        
        self.Estado_Civil_var = ctk.StringVar(value = 'Estado Civil')

        self.Estado_Civil_Menu = ctk.CTkOptionMenu(self.tabview, 
                                                   values = self.valor4, 
                                                   variable = self.Estado_Civil_var, 
                                                   command = Estado_Civil, 
                                                   width = 115, 
                                                   height = 20, 
                                                   corner_radius = 0, 
                                                   fg_color = ['#fff', '#333'], 
                                                   button_color = ['#fff', '#333'], 
                                                   button_hover_color = '', 
                                                   text_color = ['#333', '#fff'], 
                                                   font = ('arialbold', 14))
        
        def Sexo(Escolha):
            print(f'Sexo: {Escolha}')

        self.valor5 = ['Masculino', 'Feminino']
        self.Sexo_var = ctk.StringVar(value = 'Sexo')

        self.Sexo_Menu = ctk.CTkOptionMenu(self.tabview, 
                                                   values = self.valor5, 
                                                   variable = self.Sexo_var, 
                                                   command = Sexo, 
                                                   width = 115, 
                                                   height = 20, 
                                                   corner_radius = 0, 
                                                   fg_color = ['#fff', '#333'], 
                                                   button_color = ['#fff', '#333'], 
                                                   button_hover_color = '', 
                                                   text_color = ['#333', '#fff'], 
                                                   font = ('arialbold', 14))
        
    def Salvar_Dados(self):

        def Salvar():
        
                try:
                    self.ficheiro = openpyxl.load_workbook(r'Planilha admissional.xlsx', read_only = False)
                    self.folha = self.ficheiro.get_sheet_by_name(r'Dados de funcionários')
                
                except IOError:
                    messagebox.showerror(title = 'Sistema', 
                                         message = 'ERRO :/\nA planilha admissional está aberta. Feche-a antes de salvar os dados :)')
                    return
                
                self.Salvar_informações()
                self.Salvar_e_ordenar_dados()
        
        self.Botão_Salvar = ctk.CTkButton(self.tabview, 
                                          text = 'SALVAR DADOS', 
                                          width = 115, 
                                          height = 5, 
                                          command = Salvar, 
                                          font = ('arialbold', 14), 
                                          fg_color = 'green')

    def validar_data(self, data):

        padrao_data = r'\d{2}\s?/\s?\d{2}\s?/\s?\d{4}'
        return re.match(padrao_data, data) and len(re.sub(r'\D', '', data)) <= 8

    def Limpar_Informaçoes(self):

        def Limpar():
            print('Limpeza completa')
        
            self.Ficha_Entry.delete(0, END)
            self.Sap_Entry.delete(0, END)
            self.Dependentes_Entry.delete(0, END)
            self.Salario_Entry.delete(0, END)
            self.RG_Entry.delete(0, END)
            self.CPF_Entry.delete(0, END)
            self.CTPS_Entry.delete(0, END)
            self.SERIE_Entry.delete(0, END)
            self.PIS_Entry.delete(0, END)
            
            if not self.Nome_Entry.get() == '':
                self.Nome_Entry.delete(0, END)

            if not self.Contato_Entry.get() == '':
                self.Contato_Entry.delete(0, END)

            if not self.Data_Admissional_Entry.get() == '':
                self.Data_Admissional_Entry.delete(0, END)

            if not self.Data_Nasc_Entry.get() == '':
                self.Data_Nasc_Entry.delete(0, END)

        self.Botão_Limpar = ctk.CTkButton(self.tabview, 
                                          text = 'LIMPAR', 
                                          width = 115, 
                                          height = 5, 
                                          command = Limpar, 
                                          font = ('arialbold', 14), 
                                          fg_color = 'green')
        
        self.Ficha_Label.place(x = 10, y = 30)
        self.Ficha_Entry.place(x = 10, y = 55)
        
        self.Sap_Label.place(x = 10, y = 75)
        self.Sap_Entry.place(x = 10, y = 100)
        
        self.Nome_Label.place(x = 10, y = 120)
        self.Nome_Entry.place(x = 10, y = 145)

        self.Dependentes_Label.place(x = 275, y = 30)
        self.Dependentes_Entry.place(x = 275, y = 55)
        
        self.Salario_Label.place(x = 275, y = 75)
        self.Salario_Entry.place(x = 275, y = 100)

        self.Contato_Label.place(x = 372,y = 75)
        self.Contato_Entry.place(x = 372, y = 100)
        
        self.Data_Admissional_Label.place(x = 87, y = 30)
        self.Data_Admissional_Entry.place(x = 87, y = 55)

        self.Data_Nasc_Label.place(x = 87, y = 75)
        self.Data_Nasc_Entry.place(x = 87, y = 100)
        
        self.RG_Label.place(x = 275, y = 120)
        self.RG_Entry.place(x = 275, y = 145)

        self.CPF_Label.place(x = 275, y = 165)
        self.CPF_Entry.place(x = 275, y = 190)

        self.CTPS_Label.place(x = 275, y = 210)
        self.CTPS_Entry.place(x = 275, y = 235)
        
        self.SERIE_Label.place(x = 275, y = 255)
        self.SERIE_Entry.place(x = 275, y = 280)

        self.PIS_Label.place(x = 275, y = 300)
        self.PIS_Entry.place(x = 275, y = 325)
        
        self.Cargo_Menu.place(x = 10, y = 190)
        self.Departamento_Menu.place(x = 10, y = 235)
        
        self.Escolaridade_Menu.place(x = 10, y = 280)
        
        self.Estado_Civil_Menu.place(x = 10, y = 325)
        self.Sexo_Menu.place(x = 145, y = 325)

        self.Botão_Salvar.place(x = 10, y = 370)
        self.Botão_Limpar.place(x = 145, y = 370)
        
if __name__=='__main__':
    app = App()
    app.mainloop()  